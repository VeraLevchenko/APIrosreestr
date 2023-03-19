import requests
from openpyxl import load_workbook
import time

start = time.time() ## точка отсчета времени

headers = {"User-Agent": "Mozilla / 5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, как Gecko) Chrome/ 49.0.2623.110 YaBrowser / 16.4.1.8564 Safari/537.36",
		   'Content-Type': 'application/json',
		   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}

#---------------------поиск по адресу и запись в файл. Файл должен быть создан и быть пустым
wb = load_workbook('src.xlsx')
sheet = wb.active
wb_rez = load_workbook('rez1.xlsx')
sheet_rez = wb_rez.active
sheet_rez.append(('objectId', 'objectType', 'Street', 'house'))
max_row = sheet.max_row
for i in range(1, max_row):
	ul = sheet[i][0].value
	number = sheet[i][2].value
	url = 'https://rosreestr.ru/fir_lite_rest/api/gkn/address/fir_objects?macroRegionId=132000000000&regionId=132431000000&street='+ str(ul) + '&house=' + str(number)
	print(url)
	r = requests.get(url, headers=headers, verify="CertBundle.pem")
	print(ul, number)
	for a in r.json():
		print(a)
		objects = []
		objects.append(a.get("objectId"))
		objects.append(a.get("objectType"))
		objects.append(ul)
		objects.append(number)
		sheet_rez.append(objects)
		wb_rez.save('rez1.xlsx')

end = time.time() - start ## собственно время работы программы

print(end) ## вывод времени