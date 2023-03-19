import requests
from openpyxl import load_workbook
import time

start = time.time() ## точка отсчета времени

headers = {"User-Agent": "Mozilla / 5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, как Gecko) Chrome/ 49.0.2623.110 YaBrowser / 16.4.1.8564 Safari/537.36",
		   'Content-Type': 'application/json',
		   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}


#---------------------сведения ГКН по кад номеру и запись в тот же файл
wb = load_workbook('Rezult.xlsx')
sheet: object = wb.active
max_row = sheet.max_row
sheet[1][4].value = "objectType"
sheet[1][5].value = "objectName"
sheet[1][6].value = 'rightsReg'
sheet[1][7].value = 'removed'
sheet[1][8].value = 'addressNote'
for i in range(2, max_row):
	cadnum = sheet[i][0].value
	url = 'http://rosreestr.ru/fir_lite_rest/api/gkn/fir_object/' + str(cadnum)
	print(cadnum)
	r = requests.get(url, headers=headers, verify="CertBundle.pem")
	# разбираем respons ответ в формате json это словарь. из словаря достаем пару ключа parcelData, а из него rightsReg
	a = r.json().get("parcelData")
	rightsReg = a.get("rightsReg")
	# аналогично из словаря достаем пару ключа objectData, а из него removed (0 на учете, 1 снят с учета)
	b = r.json().get("objectData")
	objectType = b.get('objectType')
	objectName = b.get('objectName')
	removed = b.get('removed')
	addressNote = b.get('addressNote')
	sheet[i][4].value = objectType
	sheet[i][5].value = objectName
	sheet[i][6].value = rightsReg
	sheet[i][7].value = removed
	sheet[i][8].value = addressNote
	wb.save('Rezult.xlsx')

end = time.time() - start ## собственно время работы программы

print("Количество объектов: ", max_row)
print("Время работы кода, мин: ", end/60) ## вывод времени